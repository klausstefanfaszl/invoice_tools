#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
inbox_processor.py — Verarbeitet PDF-Rechnungsanhänge aus Exchange- oder IMAP-Postfächern.

Ablauf je Mail:
  1. Ungelesene Mails mit PDF-Anhängen werden abgerufen.
  2. Jeden PDF-Anhang temporär speichern → InvoiceExtractor ausführen.
  3. Zielverzeichnis aufbauen: <BaseDir>/<YYYY>/<MM>/
  4. Dateinamen erzeugen:      ER_MMTT_<Lieferant>_<RechnungsNr>.pdf
  5. PDF speichern, Temp-Datei löschen.
  6. Mail als gelesen markieren (nur bei fehlerfreier Verarbeitung).

Postfach-Backends:
  exchange  — Microsoft Exchange / Office 365 via EWS (exchangelib)
  imap      — Beliebige IMAP-Postfächer: Gmail, GMX, T-Online, Outlook.com, ...

Konfiguration:  invoice_inbox_config.xml  (oder per -c angeben)

Modi:
  -m dry    Simulation: zeigt was gespeichert würde, ohne Dateien zu schreiben
  -m unread Nur ungelesene Mails verarbeiten
  -m all    Alle Mails verarbeiten

Aufruf:
  python3 inbox_processor.py -m dry
  python3 inbox_processor.py -m unread -c invoice_inbox_config.xml
  python3 inbox_processor.py -m all -c andere_config.xml -d 2
"""

import argparse
import email as _email_lib
import email.header
import imaplib
import os

# Explizite Top-Level-Imports damit PyInstaller exchangelib zuverlässig einbettet
try:
    import urllib3
    import exchangelib
    import exchangelib.autodiscover
    import exchangelib.protocol
    import exchangelib.transport
    from exchangelib import Account, Configuration, Credentials, DELEGATE
    from exchangelib.protocol import BaseProtocol, NoVerifyHTTPAdapter
except ImportError:
    pass  # exchangelib ist optional (nur für Exchange-Modus benötigt)
import json
import re
import shutil

try:
    from openpyxl import load_workbook as _xl_load_workbook
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False
import subprocess
import sys
import tempfile
import unicodedata
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Callable, Dict, Iterator, List, Optional

# InvoiceExtractor aus demselben Verzeichnis laden
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from invoice_extractor import InvoiceExtractor, KiAbbruchFehler


# ── Exit-Codes ────────────────────────────────────────────────────────────────
_EXIT_OK         = 0   # Erfolgreich (inkl. keine Mails / alle übersprungen)
_EXIT_CONFIG     = 1   # Konfigurationsfehler
_EXIT_VERBINDUNG = 2   # Verbindungsfehler (Postfach nicht erreichbar)
_EXIT_KI         = 3   # KI-Abbruch (alle Provider nicht verfügbar)
_EXIT_TEILERFOLG = 4   # Teilerfolg: mind. 1 Rechnung gespeichert, mind. 1 Fehler
_EXIT_ALLE_FEHL  = 5   # Alle Anhänge fehlgeschlagen (0 gespeichert, >0 Fehler)


# ══════════════════════════════════════════════════════════════════════════════
# Gemeinsame Datenstrukturen
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class MailAttachment:
    name: str
    content: bytes


@dataclass
class MailMessage:
    subject:          str
    received:         str                              # nur für Anzeige
    attachments:      List[MailAttachment]
    _mark_read:       Callable[[], None]
    _move_to_archive: Optional[Callable[[], None]] = None

    def mark_as_read(self) -> None:
        self._mark_read()

    def move_to_archive(self) -> None:
        if self._move_to_archive:
            self._move_to_archive()


# ══════════════════════════════════════════════════════════════════════════════
# Exchange-Backend (EWS via exchangelib)
# ══════════════════════════════════════════════════════════════════════════════

def _iter_exchange(node: ET.Element, only_unread: bool) -> Iterator[MailMessage]:
    """Liefert MailMessage-Objekte aus einem Exchange-Postfach."""
    import urllib3
    from exchangelib import Account, Configuration, Credentials, DELEGATE
    from exchangelib.protocol import BaseProtocol, NoVerifyHTTPAdapter

    email_addr    = _cfg_text(node, 'Email')
    password      = _cfg_text(node, 'Password')
    server        = _cfg_text(node, 'Server')
    folder_name   = _cfg_text(node, 'Folder',        'Inbox')
    limit         = int(_cfg_text(node, 'Limit',     '100'))
    archive_name  = _cfg_text(node, 'ArchiveFolder', '')

    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    BaseProtocol.HTTP_ADAPTER_CLS = NoVerifyHTTPAdapter

    creds = Credentials(username=email_addr, password=password)
    if server:
        cfg = Configuration(server=server, credentials=creds)
        account = Account(primary_smtp_address=email_addr, credentials=creds,
                          config=cfg, autodiscover=False, access_type=DELEGATE)
    else:
        account = Account(primary_smtp_address=email_addr, credentials=creds,
                          autodiscover=True, access_type=DELEGATE)

    # Zielordner ermitteln
    target = account.inbox
    if folder_name.lower() not in ('inbox', 'posteingang', ''):
        match = next(
            (f for f in account.inbox.children if f.name.lower() == folder_name.lower()),
            None,
        )
        if match:
            target = match
        else:
            print(f'Warnung: Ordner "{folder_name}" nicht gefunden, verwende Posteingang.',
                  file=sys.stderr)

    # Archiv-Ordner suchen (Geschwister von Inbox auf Konto-Ebene)
    archive_folder = None
    if archive_name:
        try:
            archive_folder = next(
                (f for f in account.inbox.parent.children
                 if f.name.lower() == archive_name.lower()),
                None,
            )
            if archive_folder is None:
                print(f'Warnung: Exchange-Archiv-Ordner "{archive_name}" nicht gefunden.',
                      file=sys.stderr)
        except Exception as exc:
            print(f'Warnung: Archiv-Ordner konnte nicht gesucht werden: {exc}',
                  file=sys.stderr)

    query = target.all().order_by('datetime_received')
    if only_unread:
        query = query.filter(is_read=False)

    for mail in query[:limit]:
        attachments = [
            MailAttachment(name=a.name, content=a.content)
            for a in (mail.attachments or [])
            if hasattr(a, 'name') and a.name and a.name.lower().endswith('.pdf')
               and hasattr(a, 'content')
        ]
        if not attachments:
            continue

        def _mark(m=mail):
            m.is_read = True
            m.save(update_fields=['is_read'])

        def _move_exchange(m=mail, f=archive_folder):
            m.move(f)

        yield MailMessage(
            subject=mail.subject or '(kein Betreff)',
            received=str(mail.datetime_received),
            attachments=attachments,
            _mark_read=_mark,
            _move_to_archive=_move_exchange if archive_folder else None,
        )


# ══════════════════════════════════════════════════════════════════════════════
# IMAP-Backend (imaplib, Standard-Bibliothek)
# ══════════════════════════════════════════════════════════════════════════════

def _decode_mime_header(raw: Optional[str]) -> str:
    """Dekodiert MIME-Encoded-Words in Mail-Headern (z.B. =?utf-8?b?...?=)."""
    if not raw:
        return ''
    parts = email.header.decode_header(raw)
    result = []
    for part, charset in parts:
        if isinstance(part, bytes):
            result.append(part.decode(charset or 'utf-8', errors='replace'))
        else:
            result.append(part)
    return ''.join(result)


def _pdf_attachments_from_message(msg) -> List[MailAttachment]:
    """Extrahiert alle PDF-Anhänge aus einem geparsten email.message.Message-Objekt."""
    attachments = []
    for part in msg.walk():
        # Dateiname aus Content-Disposition oder Content-Type
        filename = part.get_filename()
        if not filename:
            # Fallback: name-Parameter aus Content-Type
            filename = part.get_param('name')
        if not filename:
            continue
        filename = _decode_mime_header(filename)
        if not filename.lower().endswith('.pdf'):
            continue
        payload = part.get_payload(decode=True)
        if payload:
            attachments.append(MailAttachment(name=filename, content=payload))
    return attachments


def _iter_imap(node: ET.Element, only_unread: bool) -> Iterator[MailMessage]:
    """Liefert MailMessage-Objekte aus einem IMAP-Postfach."""
    email_addr   = _cfg_text(node, 'Email')
    password     = _cfg_text(node, 'Password')
    server       = _cfg_text(node, 'Server')
    port         = int(_cfg_text(node, 'Port',         '993'))
    use_ssl      = _cfg_text(node, 'SSL',              'true').lower() == 'true'
    folder_name  = _cfg_text(node, 'Folder',           'INBOX')
    limit        = int(_cfg_text(node, 'Limit',        '100'))
    archive_name = _cfg_text(node, 'ArchiveFolder',    '')

    conn = imaplib.IMAP4_SSL(server, port) if use_ssl else imaplib.IMAP4(server, port)
    try:
        conn.login(email_addr, password)

        # Ordner öffnen
        status, detail = conn.select(f'"{folder_name}"', readonly=False)
        if status != 'OK':
            print(f'Warnung: IMAP-Ordner "{folder_name}" nicht gefunden, '
                  f'versuche INBOX.', file=sys.stderr)
            conn.select('INBOX', readonly=False)

        # UIDs suchen
        criteria  = 'UNSEEN' if only_unread else 'ALL'
        status, data = conn.uid('search', None, criteria)
        if status != 'OK' or not data or not data[0]:
            return

        uid_list = data[0].split()[:limit]   # älteste zuerst (aufsteigende UID-Reihenfolge)

        for uid in uid_list:
            status, msg_data = conn.uid('fetch', uid, '(RFC822)')
            if status != 'OK' or not msg_data or not msg_data[0]:
                continue

            raw = msg_data[0][1]
            msg = _email_lib.message_from_bytes(raw)

            attachments = _pdf_attachments_from_message(msg)
            if not attachments:
                continue

            subject  = _decode_mime_header(msg.get('Subject'))  or '(kein Betreff)'
            received = msg.get('Date', '')

            def _mark(u=uid, c=conn):
                try:
                    c.uid('store', u, '+FLAGS', '\\Seen')
                except Exception:
                    pass

            def _move_imap(u=uid, c=conn, folder=archive_name):
                try:
                    c.uid('copy', u, f'"{folder}"')
                    c.uid('store', u, '+FLAGS', '\\Deleted')
                    c.expunge()
                except Exception:
                    pass

            yield MailMessage(
                subject=subject,
                received=received,
                attachments=attachments,
                _mark_read=_mark,
                _move_to_archive=_move_imap if archive_name else None,
            )
    finally:
        try:
            conn.logout()
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════════════════
# Hilfsfunktionen
# ══════════════════════════════════════════════════════════════════════════════

def _cfg_text(root: ET.Element, path: str, default: str = '') -> str:
    el = root.find(path)
    return el.text.strip() if el is not None and el.text else default


def _sanitize(value: str, max_len: int) -> str:
    """
    Bereinigt einen String für die Verwendung als Dateinamenteil:
    Umlaute transkribieren, Unicode → ASCII, nur Alphanumerik + Bindestrich.
    """
    s = str(value)
    for src, dst in [('ä','ae'),('ö','oe'),('ü','ue'),
                     ('Ä','Ae'),('Ö','Oe'),('Ü','Ue'),('ß','ss')]:
        s = s.replace(src, dst)
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r'[^A-Za-z0-9\-]', '', s)
    return s[:max_len]


def _parse_invoice_date(date_str: str):
    """Parst DD.MM.YYYY → (day, month, year) oder ('','','') bei Fehler."""
    if date_str and re.fullmatch(r'\d{2}\.\d{2}\.\d{4}', date_str.strip()):
        day, month, year = date_str.strip().split('.')
        return day, month, year
    return '', '', ''


def _build_subdir(pattern: str, fields: dict, fallback: str) -> str:
    day, month, year = _parse_invoice_date(fields.get('InvoiceDate', ''))
    if not pattern:
        return ''
    if not year:
        return fallback
    return pattern.replace('{year}', year).replace('{month}', month)


def _build_filename(pattern: str, fields: dict,
                    supplier_max: int, invoicenr_max: int) -> str:
    day, month, year = _parse_invoice_date(fields.get('InvoiceDate', ''))
    supplier = _sanitize(fields.get('SupplierName') or 'Unbekannt', supplier_max)
    inv_nr   = _sanitize(fields.get('InvoiceNumber') or 'ohneNr',   invoicenr_max)
    name = pattern
    name = name.replace('{invoice_month}',  month)
    name = name.replace('{invoice_day}',    day)
    name = name.replace('{supplier}',       supplier)
    name = name.replace('{invoice_number}', inv_nr)
    return name


def _unique_path(path: str) -> str:
    """Hängt _2, _3 … an wenn Datei bereits existiert."""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    counter = 2
    while True:
        candidate = f'{base}_{counter}{ext}'
        if not os.path.exists(candidate):
            return candidate
        counter += 1


# ══════════════════════════════════════════════════════════════════════════════
# BankingZV-Export
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class _BzvKonto:
    """Kontodaten für die Owner-Seite (unser Bankkonto)."""
    konto_id:  str
    name:      str
    iban:      str
    acct_no:   str = ''
    bic:       str = ''
    bank_code: str = ''


@dataclass
class _BzvRegel:
    """Eine Routing-Regel: weist Rechnungen einem bestimmten Konto zu.
    Alle angegebenen Bedingungen müssen zutreffen (AND).
    Leere Bedingung = beliebig (passt immer).
    """
    account_id:       str          # Ziel-AccountId (aus <AccountMapping>)
    payment_type:     str = ''     # Substring-Match im Feld PaymentType (case-insensitiv)
    supplier_pattern: str = ''     # Regulärer Ausdruck gegen SupplierName (case-insensitiv)


@dataclass
class _BzvKfg:
    """Gesammelte BankingZV-Konfiguration aus <BankingZV> in der inbox-Config."""
    aktiv:              bool = False   # True wenn WalletPath und ExecutablePath gesetzt
    exe_pfad:           str  = r'C:\Program Files (x86)\TopBankingZV\TopBanking.exe'
    wallet_pfad:        str  = ''
    token:              str  = ''
    zahlungsziel_tage:  int  = 14
    rmtinf_template:    str  = 'Rechnung {InvoiceNumber} vom {InvoiceDate}'
    transfer_keyword:   str  = 'Überweisung'
    expected_keyword:   str  = 'Lastschrift'
    gutschrift_keyword: str  = 'Gutschrift'
    iban_fallback:      bool = True
    unattended:         bool = True
    standard_konto:     Optional[_BzvKonto] = None
    weitere_konten:     Dict[str, _BzvKonto] = field(default_factory=dict)
    routing_regeln:     List[_BzvRegel]      = field(default_factory=list)


def _lade_bzv_kfg(cfg_root: ET.Element) -> _BzvKfg:
    """Liest die <BankingZV>-Sektion aus der inbox-Config und gibt ein _BzvKfg zurück."""
    kfg  = _BzvKfg()
    node = cfg_root.find('BankingZV')
    if node is None:
        return kfg

    kfg.exe_pfad    = _cfg_text(node, 'ExecutablePath',
                                r'C:\Program Files (x86)\TopBankingZV\TopBanking.exe')
    kfg.wallet_pfad = _cfg_text(node, 'WalletPath', '')
    kfg.token       = _cfg_text(node, 'Token', '')
    term = _cfg_text(node, 'PaymentTermDays', '14')
    kfg.zahlungsziel_tage = int(term) if term.isdigit() else 14
    kfg.rmtinf_template   = _cfg_text(node, 'RmtInfTemplate',
                                       'Rechnung {InvoiceNumber} vom {InvoiceDate}')

    pt = node.find('PaymentTypeControl')
    if pt is not None:
        kfg.transfer_keyword   = _cfg_text(pt, 'TransferKeyword',       'Überweisung')
        kfg.expected_keyword   = _cfg_text(pt, 'ExpectedPaymentKeyword', 'Lastschrift')
        kfg.gutschrift_keyword = _cfg_text(pt, 'GutschriftKeyword',      'Gutschrift')
        kfg.iban_fallback      = _cfg_text(pt, 'UseFallbackIBANLogic',  'true').lower() == 'true'

    am = node.find('AccountMapping')
    if am is not None:
        dn = am.find('DefaultAccount')
        if dn is not None:
            kfg.standard_konto = _BzvKonto(
                konto_id  = _cfg_text(dn, 'AccountId',    'default'),
                name      = _cfg_text(dn, 'Name',         ''),
                iban      = _cfg_text(dn, 'AcctIBAN',     ''),
                acct_no   = _cfg_text(dn, 'AcctNo',       ''),
                bic       = _cfg_text(dn, 'AcctBIC',      ''),
                bank_code = _cfg_text(dn, 'AcctBankCode', ''),
            )
        for an in am.findall('Account'):
            kid = _cfg_text(an, 'AccountId', '')
            if kid:
                kfg.weitere_konten[kid] = _BzvKonto(
                    konto_id  = kid,
                    name      = _cfg_text(an, 'Name',         ''),
                    iban      = _cfg_text(an, 'AcctIBAN',     ''),
                    acct_no   = _cfg_text(an, 'AcctNo',       ''),
                    bic       = _cfg_text(an, 'AcctBIC',      ''),
                    bank_code = _cfg_text(an, 'AcctBankCode', ''),
                )

        for rule_el in am.findall('Routing/Rule'):
            aid = rule_el.get('accountId', '').strip()
            if aid:
                kfg.routing_regeln.append(_BzvRegel(
                    account_id       = aid,
                    payment_type     = _cfg_text(rule_el, 'PaymentType',     ''),
                    supplier_pattern = _cfg_text(rule_el, 'SupplierPattern', ''),
                ))

    kfg.aktiv = bool(kfg.wallet_pfad)
    return kfg


def _select_bzv_konto(fields: dict, kfg: '_BzvKfg') -> Optional[_BzvKonto]:
    """Wählt das BankingZV-Konto anhand der Routing-Regeln.
    Erste zutreffende Regel gewinnt; Fallback: standard_konto.
    Innerhalb einer Regel müssen alle angegebenen Bedingungen passen (AND).
    """
    payment_type = (fields.get('PaymentType') or '').strip()
    supplier     = (fields.get('SupplierName') or '').strip()

    for regel in kfg.routing_regeln:
        # PaymentType-Bedingung: Substring-Match (case-insensitiv)
        if regel.payment_type:
            if regel.payment_type.lower() not in payment_type.lower():
                continue
        # SupplierPattern-Bedingung: regulärer Ausdruck
        if regel.supplier_pattern:
            try:
                if not re.search(regel.supplier_pattern, supplier, re.IGNORECASE):
                    continue
            except re.error:
                continue
        # Alle Bedingungen erfüllt → passendes Konto suchen
        konto = kfg.weitere_konten.get(regel.account_id)
        if konto is None and kfg.standard_konto and \
                regel.account_id == kfg.standard_konto.konto_id:
            konto = kfg.standard_konto
        if konto:
            return konto

    return kfg.standard_konto


def _parse_betrag(wert: str) -> Optional[float]:
    """Parst deutschen Betrag '1.234,56 EUR' oder '1234.56' in float."""
    if not wert:
        return None
    s = re.sub(r'[^\d,.]', '', str(wert))
    # Deutsches Format: Tausenderpunkt + Komma als Dezimaltrennzeichen
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def _datum_de_zu_iso(datum: str) -> str:
    """Konvertiert DD.MM.YYYY → YYYY-MM-DD. Gibt '' bei ungültigem Format zurück."""
    if not datum:
        return ''
    m = re.fullmatch(r'(\d{2})\.(\d{2})\.(\d{4})', datum.strip())
    if m:
        return f'{m.group(3)}-{m.group(2)}-{m.group(1)}'
    return ''


def _bzv_svcLvl(fields: dict, kfg: _BzvKfg) -> str:
    """
    Bestimmt den BankingZV-SvcLvl-Wert:
      'SEPA' → Überweisung  (wir zahlen aktiv an den Lieferanten)
      'ANY'  → Erwartete Zahlung / Gutschrift (Lastschrift, Gutschrift etc.)
    """
    payment_type = (fields.get('PaymentType') or '').strip()
    if payment_type:
        if kfg.gutschrift_keyword and kfg.gutschrift_keyword.lower() in payment_type.lower():
            return 'ANY'
        if kfg.transfer_keyword and kfg.transfer_keyword.lower() in payment_type.lower():
            return 'SEPA'
        if kfg.expected_keyword and kfg.expected_keyword.lower() in payment_type.lower():
            return 'ANY'
    # Fallback: IBAN vorhanden → Überweisung möglich
    if kfg.iban_fallback:
        iban_raw = fields.get('IBAN') or ''
        if isinstance(iban_raw, list):
            iban_raw = iban_raw[0] if iban_raw else ''
        if str(iban_raw).strip():
            return 'SEPA'
    return 'ANY'


def _erstelle_bzv_eintrag(fields: dict, konto: _BzvKonto, kfg: _BzvKfg) -> dict:
    """Baut einen BankingZV-JSON-Eintrag aus den extrahierten Rechnungsfeldern."""
    inv_nr   = str(fields.get('InvoiceNumber') or '')
    inv_date = str(fields.get('InvoiceDate')   or '')
    supplier = str(fields.get('SupplierName')  or '')
    gross    = _parse_betrag(str(fields.get('GrossAmount') or ''))

    # IBAN (kann Liste sein bei multi="true" in der Extractor-Config)
    iban_raw = fields.get('IBAN') or ''
    if isinstance(iban_raw, list):
        iban_raw = iban_raw[0] if iban_raw else ''
    iban_raw = str(iban_raw).strip()

    # Fälligkeitsdatum: DueDate aus Extraktion, sonst Rechnungsdatum + Zahlungsziel
    due_iso = _datum_de_zu_iso(str(fields.get('DueDate') or ''))
    if not due_iso:
        inv_iso = _datum_de_zu_iso(inv_date)
        if inv_iso:
            try:
                due_iso = (date.fromisoformat(inv_iso)
                           + timedelta(days=kfg.zahlungsziel_tage)).isoformat()
            except ValueError:
                due_iso = ''

    # Verwendungszweck aus Template
    rmtinf = kfg.rmtinf_template
    rmtinf = rmtinf.replace('{InvoiceNumber}', inv_nr)
    rmtinf = rmtinf.replace('{InvoiceDate}',   inv_date)
    rmtinf = rmtinf.replace('{SupplierName}',  supplier)

    svc_lvl = _bzv_svcLvl(fields, kfg)
    payment_type = (fields.get('PaymentType') or '').strip()
    # Gutschrift (credit note): SvcLvl=ANY + PmtMtd=DD → "Erwartete Gutschrift" in BankingZV
    # Lastschrift/andere ANY: SvcLvl=ANY + PmtMtd=TRF → "Erwartete Zahlung"
    # Überweisung: SvcLvl=SEPA + PmtMtd=TRF → Überweisung
    if svc_lvl == 'ANY' and kfg.gutschrift_keyword and \
            kfg.gutschrift_keyword.lower() in payment_type.lower():
        pmt_mtd = 'DD'
    else:
        pmt_mtd = 'TRF'

    return {
        'Id':               inv_nr,
        'SvcLvl':           svc_lvl,
        'PmtMtd':           pmt_mtd,
        'ReqdExctnDt':      due_iso,
        'Amt':              f'{gross:.2f}' if gross is not None else '0.00',
        'AmtCcy':           'EUR',
        'EndToEndId':       inv_nr,
        'RmtInf':           rmtinf,
        'OwnrNm':           konto.name,
        'OwnrAcctCtry':     'DE',
        'OwnrAcctIBAN':     konto.iban,
        'OwnrAcctNo':       konto.acct_no,
        'OwnrAcctBIC':      konto.bic,
        'OwnrAcctBankCode': konto.bank_code,
        'RmtdNm':           supplier,
        'RmtdAcctCtry':     'DE',
        'RmtdAcctIBAN':     iban_raw,
    }


_BZV_EXITCODES: Dict[int, str] = {
    1000:  'Benutzer hat Kommando abgebrochen (z.B. PIN-Eingabe)',
    9000:  'Bank hat bei der Übertragung einen Fehlercode gesendet',
    10000: 'Ungültige Kommandozeilenparameter',
    10001: 'Datentresor konnte nicht selektiert werden (Pfad prüfen)',
    10002: 'Datentresor konnte nicht geöffnet werden (Token/Passwort falsch?)',
    10003: 'Datentresor existiert nicht und konnte nicht angelegt werden',
    10004: 'Importdatei war ungültig (JSON-Format prüfen)',
    10005: 'Keine Business-Lizenz — Kommandozeilenmodus erfordert Business-Lizenz',
    20000: 'Interner Fehler in BankingZV',
}


def _rufe_bankingzv(kfg: _BzvKfg, acct_iban: str, json_pfad: str) -> int:
    """Ruft TopBanking.exe über die CLI-Schnittstelle auf. Gibt Exit-Code zurück."""
    cmd = [kfg.exe_pfad, '-Cmd',
           '-Wallet',      kfg.wallet_pfad,
           '-Token',       kfg.token,
           '-AcctIBAN',    acct_iban,
           '-ImportPaymts', json_pfad]
    if kfg.unattended:
        cmd.append('-Unattended')
    try:
        return subprocess.run(cmd, capture_output=True, text=True).returncode
    except FileNotFoundError:
        return -1
    except Exception:
        return -99


def _exportiere_zu_bankingzv(eintraege: List[dict], kfg: _BzvKfg,
                              bzv_modus: str, mail_dry_run: bool) -> None:
    """Exportiert Einträge zu BankingZV gemäß bzv_modus:
       'dry'    – nur Anzeige, keine JSON-Datei, kein BankingZV-Aufruf
       'json'   – Anzeige + JSON-Datei bleibt liegen, kein BankingZV-Aufruf
       'export' – Anzeige + JSON-Datei + BankingZV-Aufruf
       (bei mail_dry_run wird 'export' wie 'json' behandelt)
    """
    if not bzv_modus:
        return

    if not eintraege:
        print('\nBankingZV: Keine Einträge zum Exportieren.')
        return

    konto = kfg.standard_konto
    if not konto:
        print('\nBankingZV: Kein Konto konfiguriert, Export übersprungen.', file=sys.stderr)
        return

    # ── Zusammenfassung anzeigen ──────────────────────────────────────────────
    dry_prefix = '[DRY-RUN] ' if mail_dry_run else ''
    modus_hint = f' [{bzv_modus}]' if bzv_modus != 'export' else ''
    print(f'\n{dry_prefix}BankingZV-Export{modus_hint}: {len(eintraege)} Eintrag/Einträge'
          f' -> Konto {konto.konto_id} ({konto.iban})')

    ueberweisung = [e for e in eintraege if e['SvcLvl'] == 'SEPA']
    gutschriften  = [e for e in eintraege if e['SvcLvl'] == 'ANY' and e.get('PmtMtd') == 'DD']
    zahlungen     = [e for e in eintraege if e['SvcLvl'] == 'ANY' and e.get('PmtMtd') != 'DD']
    if ueberweisung:
        print(f'  Überweisungen      : {len(ueberweisung)}')
    if zahlungen:
        print(f'  Erwartete Zahlungen: {len(zahlungen)}')
    if gutschriften:
        print(f'  Erw. Gutschriften  : {len(gutschriften)}')

    for e in eintraege:
        if e['SvcLvl'] == 'SEPA':
            typ = 'Überweisung   '
        elif e.get('PmtMtd') == 'DD':
            typ = 'Erw. Gutschrift'
        else:
            typ = 'Erw. Zahlung  '
        print(f'  {typ}  {e["Amt"]:>10s} EUR  {e["RmtdNm"]}  [{e["Id"]}]')

    if bzv_modus == 'dry':
        return  # nur Anzeige — fertig

    # ── JSON-Datei schreiben ──────────────────────────────────────────────────
    tmp_fd, json_pfad = tempfile.mkstemp(suffix='_2bZV.json', prefix='inbox_')
    loesche_json = True
    try:
        with os.fdopen(tmp_fd, 'w', encoding='utf-8') as fh:
            json.dump(eintraege, fh, ensure_ascii=False, indent=2)
        print(f'  JSON: {json_pfad}')

        if bzv_modus == 'json' or mail_dry_run:
            loesche_json = False  # JSON bleibt zur Kontrolle liegen
            if mail_dry_run:
                print(f'  [DRY-RUN] BankingZV-Aufruf unterdrückt.')
            return

        # ── BankingZV aufrufen (nur bei 'export' und kein mail_dry_run) ───────
        if not os.path.isfile(kfg.exe_pfad):
            print(f'  Warnung: TopBanking.exe nicht gefunden: {kfg.exe_pfad}',
                  file=sys.stderr)
            loesche_json = False  # JSON bleibt zur manuellen Verwendung
            return

        exit_code = _rufe_bankingzv(kfg, konto.iban, json_pfad)
        if exit_code == 0:
            print(f'  BankingZV: Import erfolgreich.')
        else:
            msg = _BZV_EXITCODES.get(exit_code, 'Unbekannter Fehler')
            print(f'  BankingZV: Fehler (Exit {exit_code}): {msg}', file=sys.stderr)
            loesche_json = False  # bei Fehler JSON behalten
    finally:
        if loesche_json and os.path.exists(json_pfad):
            os.unlink(json_pfad)


# ══════════════════════════════════════════════════════════════════════════════
# Excel-Export (Rechnungseingang)
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class _ExcelSpalte:
    """Eine Zeile im Spaltenmapping: Excel-Spaltenname → Quelle + Typ.

    quelle:  Extraktionsfeldname (z.B. 'InvoiceDate') oder Sonderwert:
               {StandardTyp}  → kfg.standard_typ
               {KontoId}      → konto_id-Parameter (aus BankingZV-Routing)
    typ:     Konvertierungstyp:
               'datum'   → date-Objekt  (aus DD.MM.YYYY)
               'betrag'  → float        (aus deutschem Betrag)
               'iban'    → str, erstes Element wenn Liste
               ''        → str (kein Leerstring → None)
    """
    name:   str        # Tabellenspaltennamen in Excel
    quelle: str        # Feldname oder {Sonderwert}
    typ:    str = ''


# Standardmapping — wird verwendet wenn kein <Spaltenmapping> in der Config
_EXCEL_DEFAULT_MAPPING: List[_ExcelSpalte] = [
    _ExcelSpalte('Re-Datum',       'InvoiceDate',   'datum'),
    _ExcelSpalte('Name/Lieferant', 'SupplierName',  ''),
    _ExcelSpalte('RE-Nr',          'InvoiceNumber', ''),
    _ExcelSpalte('TYP',            '{StandardTyp}', ''),
    _ExcelSpalte('Netto',          'NetAmount',     'betrag'),
    _ExcelSpalte('Brutto',         'GrossAmount',   'betrag'),
    _ExcelSpalte('Fällig_am',      'DueDate',       'datum'),
    _ExcelSpalte('IBAN',           'IBAN',          'iban'),
    _ExcelSpalte('Konto-ID',       '{KontoId}',     ''),
]


@dataclass
class _ExcelKfg:
    """Konfiguration für den Excel-Export in die Rechnungseingangs-Tabelle."""
    aktiv:          bool               = False
    datei_pfad:     str                = ''
    tabellenblatt:  str                = 'ER'
    tabellenname:   str                = 'tb_rechnungen'
    standard_typ:   str                = 'E'
    duplikat_spalte: str               = 'RE-Nr'
    spaltenmapping: List[_ExcelSpalte] = field(default_factory=lambda: list(_EXCEL_DEFAULT_MAPPING))


def _lade_excel_kfg(cfg_root: ET.Element) -> '_ExcelKfg':
    """Liest die <ExcelExport>-Sektion aus der inbox-Config."""
    kfg  = _ExcelKfg()
    node = cfg_root.find('ExcelExport')
    if node is None:
        return kfg
    kfg.datei_pfad      = _cfg_text(node, 'DateiPfad',      '')
    kfg.tabellenblatt   = _cfg_text(node, 'Tabellenblatt',  'ER')
    kfg.tabellenname    = _cfg_text(node, 'Tabellenname',   'tb_rechnungen')
    kfg.standard_typ    = _cfg_text(node, 'StandardTyp',    'E')
    kfg.duplikat_spalte = _cfg_text(node, 'DuplikatSpalte', 'RE-Nr')

    mapping_node = node.find('Spaltenmapping')
    if mapping_node is not None:
        mapping: List[_ExcelSpalte] = []
        for el in mapping_node.findall('Spalte'):
            spalten_name = (el.get('name') or '').strip()
            quelle       = (el.text        or '').strip()
            typ          = (el.get('typ')  or '').strip().lower()
            if spalten_name and quelle:
                mapping.append(_ExcelSpalte(spalten_name, quelle, typ))
        if mapping:
            kfg.spaltenmapping = mapping

    kfg.aktiv = bool(kfg.datei_pfad)
    return kfg


def _schreibe_in_excel(fields: dict, konto_id: str,
                        kfg: '_ExcelKfg', dry_run: bool,
                        debug: int = 0) -> bool:
    """Fügt eine neue Zeile mit Rechnungsdaten in die Excel-Tabelle ein.

    Die Zielspalten werden über den Namen der Excel-Tabelle (<Tabellenname>)
    und deren Spaltenköpfe bestimmt — keine hardcodierten Spaltennummern.
    Formel-Spalten (calculatedColumnFormula) werden aus der letzten Zeile
    kopiert. Die Tabelle wird auf die neue Zeile ausgedehnt.

    Vor dem Schreiben wird die RE-Nr-Spalte auf Duplikate geprüft.
    Ein bereits vorhandener Eintrag erzeugt einen Info-Hinweis (kein Fehler).

    Gibt True bei Erfolg oder Duplikat zurück, False bei Fehler.
    """
    if not _OPENPYXL_OK:
        print('  Warnung: openpyxl nicht installiert — Excel-Export übersprungen.',
              file=sys.stderr)
        return False

    if not kfg.datei_pfad or not os.path.isfile(kfg.datei_pfad):
        print(f'  Warnung: Excel-Datei nicht gefunden: {kfg.datei_pfad}',
              file=sys.stderr)
        return False

    # ── Werte aus Spaltenmapping auflösen ─────────────────────────────────────
    def _str_zu_date(s: str) -> Optional[date]:
        d, m, y = _parse_invoice_date(s or '')
        if y:
            try:
                return date(int(y), int(m), int(d))
            except ValueError:
                pass
        return None

    def _wert(sp: '_ExcelSpalte') -> object:
        """Liefert den aufgelösten Zellwert für eine Mapping-Zeile."""
        if sp.quelle == '{StandardTyp}':
            raw = kfg.standard_typ or None
        elif sp.quelle == '{KontoId}':
            raw = konto_id or None
        else:
            raw = fields.get(sp.quelle)
        if raw is None:
            return None
        if sp.typ == 'datum':
            return _str_zu_date(str(raw))
        if sp.typ == 'betrag':
            return _parse_betrag(str(raw))
        if sp.typ == 'iban':
            if isinstance(raw, list):
                raw = raw[0] if raw else ''
            return str(raw).strip() or None
        # Standardtyp: String
        val = str(raw).strip() or None
        # Rechnungsnummern die nur aus Ziffern bestehen mit "R" präfixen,
        # damit sie in Excel als Text erkennbar bleiben und beim erneuten
        # Einlesen trotzdem als Duplikat erkannt werden (Substring-Match).
        if val and sp.quelle == 'InvoiceNumber' and val.isdigit():
            val = 'R' + val
        return val

    # Spaltenname → Wert (nur nicht-None-Einträge)
    werte: Dict[str, object] = {
        sp.name: _wert(sp)
        for sp in kfg.spaltenmapping
    }

    # inv_nr separat für Duplikatprüfung bereitstellen
    inv_nr = str(fields.get('InvoiceNumber') or '').strip()

    # ── Workbook öffnen (auch im Dry-Run, für Duplikat-Prüfung) ──────────────
    try:
        import warnings as _warnings
        with _warnings.catch_warnings():
            _warnings.simplefilter('ignore', UserWarning)
            wb = _xl_load_workbook(kfg.datei_pfad)
    except Exception as exc:
        print(f'  Warnung: Excel-Datei konnte nicht geöffnet werden: {exc}',
              file=sys.stderr)
        return False

    try:
        if kfg.tabellenblatt not in wb.sheetnames:
            print(f'  Warnung: Tabellenblatt "{kfg.tabellenblatt}" nicht gefunden.',
                  file=sys.stderr)
            return False

        ws = wb[kfg.tabellenblatt]

        # ── Tabelle und Spalten-Mapping aufbauen ──────────────────────────────
        if kfg.tabellenname not in ws.tables:
            print(f'  Warnung: Tabelle "{kfg.tabellenname}" nicht in '
                  f'Blatt "{kfg.tabellenblatt}" gefunden.', file=sys.stderr)
            return False

        tbl     = ws.tables[kfg.tabellenname]
        ref_m   = re.fullmatch(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', tbl.ref)
        if not ref_m:
            print(f'  Warnung: Tabellenreferenz "{tbl.ref}" nicht lesbar.',
                  file=sys.stderr)
            return False

        from openpyxl.utils import column_index_from_string
        start_col  = column_index_from_string(ref_m.group(1))
        end_col_l  = ref_m.group(3)
        last_row   = int(ref_m.group(4))

        # Spaltenname → absolute Spaltennummer + Formel-Spalten erkennen
        col_map:     Dict[str, int]  = {}
        formel_cols: List[int]       = []
        for i, tc in enumerate(tbl.tableColumns):
            abs_col = start_col + i
            col_map[tc.name] = abs_col
            if tc.calculatedColumnFormula is not None:
                formel_cols.append(abs_col)

        # ── Duplikat-Prüfung (unscharfe RE-Nr + Datum/Betrag/Name) ──────────
        dup_col = col_map.get(kfg.duplikat_spalte)
        dup_val = str(werte.get(kfg.duplikat_spalte) or inv_nr).strip()
        if dup_val and dup_col:

            def _re_nr_aehnlich(a: str, b: str) -> bool:
                """True wenn a und b als gleiche RE-Nr gelten.
                Extra Zeichen am Anfang/Ende (z.B. Präfix 'RE-') werden toleriert."""
                a, b = a.lower().strip(), b.lower().strip()
                return a == b or a in b or b in a

            def _name_aehnlich(a: str, b: str) -> bool:
                if not a or not b:
                    return True   # fehlender Name schließt Duplikat nicht aus
                a, b = a.lower(), b.lower()
                if a in b or b in a:
                    return True
                from difflib import SequenceMatcher
                return SequenceMatcher(None, a, b).ratio() >= 0.70

            def _betrag_gleich(v1, v2) -> bool:
                try:
                    return float(v1) == float(v2)
                except (TypeError, ValueError):
                    return False

            # Excel-Spaltenindizes für Sekundär-Kriterien ermitteln
            def _quelle_col(quelle: str) -> Optional[int]:
                for sp in kfg.spaltenmapping:
                    if sp.quelle == quelle and sp.name in col_map:
                        return col_map[sp.name]
                return None

            datum_col_idx  = _quelle_col('InvoiceDate')
            betrag_col_idx = _quelle_col('GrossAmount')
            name_col_idx   = _quelle_col('SupplierName')

            # Neue Werte für Sekundär-Vergleich
            neu_datum  = next(
                (werte[sp.name] for sp in kfg.spaltenmapping
                 if sp.quelle == 'InvoiceDate' and sp.name in werte), None)
            neu_betrag = next(
                (werte[sp.name] for sp in kfg.spaltenmapping
                 if sp.quelle == 'GrossAmount' and sp.name in werte), None)
            neu_name   = str(fields.get('SupplierName') or '').strip()

            dup_norm   = dup_val.lower()
            header_row = int(ref_m.group(2))

            for row_idx in range(header_row + 1, last_row + 1):
                zell_re_nr = str(ws.cell(row_idx, dup_col).value or '').strip()
                if not zell_re_nr or not _re_nr_aehnlich(dup_norm, zell_re_nr):
                    continue

                # RE-Nr ähnlich → Datum, Betrag und Name als entscheidende Kriterien
                datum_ok = betrag_ok = name_ok = True

                if datum_col_idx is not None and neu_datum is not None:
                    zell_datum_raw = ws.cell(row_idx, datum_col_idx).value
                    # openpyxl liefert datetime.datetime; auf date normalisieren
                    if hasattr(zell_datum_raw, 'date'):
                        zell_datum_raw = zell_datum_raw.date()
                    datum_ok = (zell_datum_raw == neu_datum)

                if betrag_col_idx is not None and neu_betrag is not None:
                    betrag_ok = _betrag_gleich(
                        ws.cell(row_idx, betrag_col_idx).value, neu_betrag)

                if name_col_idx is not None and neu_name:
                    zell_name = str(ws.cell(row_idx, name_col_idx).value or '').strip()
                    name_ok = _name_aehnlich(neu_name, zell_name)

                if datum_ok and betrag_ok and name_ok:
                    print(f'    Info: Excel-Export übersprungen — '
                          f'Duplikat in Zeile {row_idx} erkannt '
                          f'(RE-Nr "{zell_re_nr}" ~ "{dup_val}", '
                          f'Datum/Betrag/Name stimmen überein).')
                    return True  # kein Fehler

        # ── Dry-Run-Vorschau ──────────────────────────────────────────────────
        if dry_run:
            print(f'    [DRY-RUN] Excel -> {kfg.datei_pfad} '
                  f'[{kfg.tabellenblatt}/{kfg.tabellenname}]')
            if debug > 0:
                for col_name, wert in werte.items():
                    if wert is not None and col_name in col_map:
                        print(f'      {col_name:<16}: {wert}')
                if formel_cols:
                    namen = [n for n, c in col_map.items() if c in formel_cols]
                    print(f'      Formel-Spalten  : {", ".join(namen)} '
                          f'(aus Vorgaengerzeile)')
            return True

        # ── Neue Zeile schreiben ──────────────────────────────────────────────
        # Erste leere Zeile innerhalb der Tabelle suchen.
        # Nur Datenspalten prüfen (Formel-Spalten werden bewusst ausgeschlossen,
        # da sie auch in leeren Zeilen Werte enthalten können).
        header_row_idx = int(ref_m.group(2))
        daten_cols     = [c for c in col_map.values() if c not in formel_cols]
        next_row       = None
        for r in range(header_row_idx + 1, last_row + 1):
            if daten_cols and all(ws.cell(r, c).value in (None, '')
                                  for c in daten_cols):
                next_row = r
                break

        extend_tbl = next_row is None
        if extend_tbl:
            next_row = last_row + 1

        # ── Formatierung aus Vorgängerzeile sicherstellen ─────────────────────
        # Bei neuen Zeilen (extend_tbl): komplette Formatierung übernehmen.
        # Bei vorhandenen leeren Zeilen: nur number_format, falls noch nicht
        # gesetzt (damit Datum- und Betragsfelder korrekt dargestellt werden).
        prev_row = next_row - 1
        end_col_idx = column_index_from_string(end_col_l)
        from copy import copy as _copy
        for col_idx in range(start_col, end_col_idx + 1):
            if col_idx in formel_cols:
                continue  # Formel-Spalten nicht anfassen — Formatierung bleibt unverändert
            src = ws.cell(prev_row, col_idx)
            dst = ws.cell(next_row, col_idx)
            if extend_tbl:
                dst.font          = _copy(src.font)
                dst.border        = _copy(src.border)
                dst.fill          = _copy(src.fill)
                dst.number_format = src.number_format
                dst.protection    = _copy(src.protection)
                dst.alignment     = _copy(src.alignment)
            elif dst.number_format in ('General', 'general', '', None):
                dst.number_format = src.number_format

        for col_name, wert in werte.items():
            if wert is not None and col_name in col_map:
                ws.cell(next_row, col_map[col_name], wert)

        # ── Formeln aus Vorgaengerzeile kopieren ──────────────────────────────
        for col_idx in formel_cols:
            prev_cell = ws.cell(prev_row, col_idx)
            if prev_cell.value and str(prev_cell.value).startswith('='):
                ws.cell(next_row, col_idx, prev_cell.value)

        # ── Tabellenreferenz ausdehnen (nur wenn Zeile außerhalb der Tabelle) ─
        if extend_tbl:
            tbl.ref = f'{ref_m.group(1)}{ref_m.group(2)}:{end_col_l}{next_row}'

        wb.save(kfg.datei_pfad)
        _datum   = str(fields.get('InvoiceDate')  or 'kein Datum')
        _liefer  = str(fields.get('SupplierName') or '?')
        print(f'    Excel [{kfg.tabellenname}]: Zeile {next_row} eingefuegt '
              f'({_datum}, {_liefer}, {inv_nr or "?"})')
        return True

    except Exception as exc:
        print(f'  Warnung: Excel-Export fehlgeschlagen: {exc}', file=sys.stderr)
        return False


# ══════════════════════════════════════════════════════════════════════════════
# Hauptprogramm
# ══════════════════════════════════════════════════════════════════════════════

def run(argv=None):
    parser = argparse.ArgumentParser(
        description='Verarbeitet PDF-Rechnungsanhänge aus Exchange- oder IMAP-Postfächern.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            'Modi:\n'
            '  dry    Simulation: zeigt was gespeichert würde, ohne Dateien zu schreiben\n'
            '         (nur ungelesene Mails; mit -e wird Excel trotzdem geschrieben)\n'
            '  dryall Wie dry, aber alle Mails — auch bereits gelesene;\n'
            '         nützlich um ältere Rechnungen nachträglich per -e in Excel zu übertragen\n'
            '  unread Nur ungelesene Mails verarbeiten und speichern\n'
            '  all    Alle Mails verarbeiten und speichern\n'
            '  archiv Wie unread, verschiebt erfolgreich verarbeitete Mails zusätzlich\n'
            '         in den Archiv-Ordner (<ArchiveFolder> in Config, Standard: "Archiv")\n\n'
            'Beispiele:\n'
            '  python3 inbox_processor.py -m dry\n'
            '  python3 inbox_processor.py -m unread -c invoice_inbox_config.xml\n'
            '  python3 inbox_processor.py -m archiv -c invoice_inbox_config.xml\n'
            '  python3 inbox_processor.py -m all -c andere_config.xml -d 2\n'
            '  python3 inbox_processor.py -m unread -e\n'
            '  python3 inbox_processor.py -m unread -e -b export\n'
            '  python3 inbox_processor.py -m dry -B "F:\\Buchhaltung"\n'
            '  python3 inbox_processor.py -m unread -B "F:\\Buchhaltung" -e\n\n'
            'Excel-Export (-e / --export-excel):\n'
            '  Schreibt jede erkannte Rechnung als neue Zeile in die Tabelle\n'
            '  <ExcelExport><DateiPfad> aus der Config. Vor dem Schreiben wird\n'
            '  geprüft ob die RE-Nr bereits vorhanden ist (kein Duplikat möglich).\n'
            '  Pfad zur Excel-Datei wird in <ExcelExport><DateiPfad> konfiguriert.'
        ),
    )
    parser.add_argument('-c', '--config', default='invoice_inbox_config.xml',
                        metavar='CONFIGDATEI',
                        help='XML-Konfigurationsdatei (Standard: invoice_inbox_config.xml)')
    parser.add_argument('-m', '--modus', required=True,
                        choices=['dry', 'dryall', 'unread', 'all', 'archiv'], metavar='MODUS',
                        help='dry=Simulation (ungelesene) | dryall=Simulation (alle Mails) | '
                             'unread=nur ungelesene | all=alle | '
                             'archiv=ungelesene + Archivierung nach Verarbeitung')
    parser.add_argument('--dry-run', action='store_true', default=False,
                        help='Simulation: keine Dateien speichern, kein BankingZV-Aufruf '
                             '(kombinierbar mit -m all oder -m unread); '
                             'mit -e wird Excel trotzdem tatsächlich geschrieben')
    parser.add_argument('-d', '--debug', type=int, default=0, metavar='LEVEL',
                        help='Debug-Level für InvoiceExtractor (0=aus … 3=Vollausgabe)')
    parser.add_argument('-a', '--api', default=None, metavar='API_CONFIG',
                        help='Zentrale KI-API-Konfigurationsdatei '
                             '(Standard: invoice_tools_api_config.xml neben der Config)')
    parser.add_argument('-b', '--bzv', choices=['dry', 'json', 'export'], default='',
                        metavar='MODUS',
                        help='BankingZV-Export aktivieren: '
                             'dry=nur Anzeige | json=Anzeige+JSON-Datei | export=+BankingZV-Aufruf. '
                             'Wallet und Token kommen aus invoice_inbox_config.xml.')
    parser.add_argument('-B', '--bdir', default=None, metavar='VERZEICHNIS',
                        help='Basisverzeichnis für die PDF-Ablage; überschreibt '
                             '<BaseDir> aus der Config-Datei.')
    parser.add_argument('-e', '--export-excel', action='store_true', default=False,
                        help='Rechnungsdaten in die Excel-Eingangstabelle schreiben '
                             '(Pfad wird aus <ExcelExport><DateiPfad> in der Config gelesen).')
    parser.add_argument('-l', '--log', default=None, metavar='LOGDATEI',
                        help='Protokolldatei; ohne -d wird stdout vollständig dorthin '
                             'umgeleitet (kein Ausgabe auf stdout im Normalbetrieb)')
    if argv is not None and len(argv) == 0:
        parser.print_help()
        sys.exit(0)

    args = parser.parse_args(argv)

    # ── Log-Datei / stdout-Umleitung ─────────────────────────────────────────
    _log_fh      = None
    _orig_stdout = sys.stdout
    if args.log:
        try:
            _log_fh = open(os.path.abspath(args.log), 'w', encoding='utf-8')
            print(f'# inbox_processor — {datetime.now():%Y-%m-%d %H:%M:%S}'
                  f' — Modus: {args.modus}', file=_log_fh)
            if args.debug == 0:
                sys.stdout = _log_fh    # alle print() → Log, stdout stumm
        except OSError as e:
            print(f'Warnung: Log-Datei konnte nicht geöffnet werden: {e}',
                  file=sys.stderr)
            _log_fh = None

    # ── Config laden ──────────────────────────────────────────────────────────
    config_path = os.path.abspath(args.config)
    if not os.path.isfile(config_path):
        print(f'Fehler: Konfigurationsdatei nicht gefunden: {config_path}', file=sys.stderr)
        sys.exit(1)

    cfg        = ET.parse(config_path).getroot()
    config_dir = os.path.dirname(config_path)

    # ── BankingZV-Konfiguration ───────────────────────────────────────────────
    bzv_kfg      = _lade_bzv_kfg(cfg)
    bzv_modus    = args.bzv        # '' | 'dry' | 'json' | 'export'
    bzv_eintraege: List[dict] = []
    if bzv_modus and bzv_kfg.aktiv:
        print(f'BankingZV [{bzv_modus}]: Wallet={bzv_kfg.wallet_pfad}'
              + (f', Konto={bzv_kfg.standard_konto.iban}' if bzv_kfg.standard_konto else ''))
    elif bzv_modus and not bzv_kfg.aktiv:
        print(f'Warnung: BankingZV-Modus "{bzv_modus}" angefordert, '
              f'aber kein WalletPath in der Config konfiguriert — BankingZV ignoriert.',
              file=sys.stderr)
        bzv_modus = ''  # deaktivieren

    # ── Excel-Export-Konfiguration ────────────────────────────────────────────
    excel_kfg    = _lade_excel_kfg(cfg)
    excel_export = args.export_excel
    if excel_export and not excel_kfg.aktiv:
        print(f'Warnung: --export-excel angefordert, aber kein <DateiPfad> in '
              f'<ExcelExport> konfiguriert — Excel-Export ignoriert.', file=sys.stderr)
        excel_export = False
    elif excel_export:
        print(f'Excel-Export: {excel_kfg.datei_pfad} [{excel_kfg.tabellenblatt}]')

    # Postfach-Konfiguration
    mailbox_node = cfg.find('Mailbox')
    if mailbox_node is None:
        print('Fehler: Kein <Mailbox>-Element in der Konfiguration gefunden.', file=sys.stderr)
        sys.exit(1)

    mailbox_type = (mailbox_node.get('type') or '').lower()
    if mailbox_type not in ('exchange', 'imap'):
        print(f'Fehler: Unbekannter Postfach-Typ "{mailbox_type}". '
              f'Erlaubt: exchange, imap', file=sys.stderr)
        sys.exit(1)

    email_addr = _cfg_text(mailbox_node, 'Email')
    password   = _cfg_text(mailbox_node, 'Password')
    if not email_addr or not password:
        print('Fehler: <Email> und <Password> sind Pflichtfelder in <Mailbox>.',
              file=sys.stderr)
        sys.exit(1)

    mark_as_read = _cfg_text(mailbox_node, 'MarkAsRead', 'true').lower() == 'true'

    # IMAP: Server ist Pflicht
    if mailbox_type == 'imap' and not _cfg_text(mailbox_node, 'Server'):
        print('Fehler: <Server> ist bei type="imap" ein Pflichtfeld.', file=sys.stderr)
        sys.exit(1)

    # Anhang-Filter — Dateinamen-Muster zum Überspringen
    skip_patterns = [el.text.strip().lower()
                     for el in cfg.findall('AttachmentFilter/SkipPattern')
                     if el.text and el.text.strip()]

    # Rechnungsfilter — Pflichtfelder
    required_fields = [el.text.strip()
                       for el in cfg.findall('InvoiceFilter/RequiredField')
                       if el.text and el.text.strip()]

    # Ablage
    base_dir        = _cfg_text(cfg, 'Storage/BaseDir', '.')
    if args.bdir:
        base_dir = args.bdir
        print(f'BaseDir (überschrieben): {base_dir}')
    subpath_pattern = _cfg_text(cfg, 'Storage/Subpath', '{year}/{month}')
    fallback_dir    = _cfg_text(cfg, 'Storage/FallbackDir', '_unbekannt')

    # Dateiname
    filename_pattern = _cfg_text(cfg, 'Filename/Pattern',
                                  'ER_{invoice_month}{invoice_day}_{supplier}_{invoice_number}')
    supplier_max  = int(_cfg_text(cfg, 'Filename/SupplierMaxLen',     '20'))
    invoicenr_max = int(_cfg_text(cfg, 'Filename/InvoiceNumberMaxLen','20'))

    # InvoiceExtractor
    extractor_config = _cfg_text(cfg, 'InvoiceExtractor/Config',
                                  'invoice_extractor_config_RE.xml')
    if not os.path.isabs(extractor_config):
        extractor_config = os.path.join(config_dir, extractor_config)
    if not os.path.isfile(extractor_config):
        print(f'Fehler: InvoiceExtractor-Config nicht gefunden: {extractor_config}',
              file=sys.stderr)
        sys.exit(1)

    # ── Modus auswerten ───────────────────────────────────────────────────────
    dry_run     = args.modus in ('dry', 'dryall') or args.dry_run
    only_unread = args.modus in ('dry', 'unread', 'archiv')

    # Bei Modus "archiv": ArchiveFolder sicherstellen (Config-Wert oder Default "Archiv")
    if args.modus == 'archiv':
        af_el = mailbox_node.find('ArchiveFolder')
        if af_el is None:
            af_el = ET.SubElement(mailbox_node, 'ArchiveFolder')
        if not (af_el.text and af_el.text.strip()):
            af_el.text = 'Archiv'

    # ── Verbindung aufbauen und Mails verarbeiten ─────────────────────────────
    print(f'Verbinde mit {mailbox_type.upper()}-Postfach ({email_addr})...')
    print(f'Modus: {args.modus}')

    extractor  = InvoiceExtractor(extractor_config, debug_level=args.debug,
                                   api_config_path=args.api)
    dry_prefix       = '[DRY-RUN] ' if dry_run else ''
    count_ok         = 0
    count_err        = 0
    _verbindungsfehler = False

    try:
        if mailbox_type == 'exchange':
            mail_iter = _iter_exchange(mailbox_node, only_unread)
        else:
            mail_iter = _iter_imap(mailbox_node, only_unread)

        for mail in mail_iter:
            print(f'\nMail: {mail.subject}  ({mail.received})')
            mail_had_error = False

            for attachment in mail.attachments:
                print(f'  Anhang: {attachment.name}')

                # Dateinamen-Filter prüfen
                name_lower = attachment.name.lower()
                matched = next((p for p in skip_patterns if p in name_lower), None)
                if matched:
                    print(f'  Übersprungen (Dateiname enthält "{matched}")')
                    continue

                tmp_fd, tmp_path = tempfile.mkstemp(suffix='.pdf')
                try:
                    with os.fdopen(tmp_fd, 'wb') as fh:
                        fh.write(attachment.content)

                    try:
                        fields = extractor.extract(tmp_path)
                    except KiAbbruchFehler as e:
                        print(f'\nFEHLER: {e}', file=sys.stderr)
                        print('Verarbeitung abgebrochen — KI-API nicht verfügbar.',
                              file=sys.stderr)
                        print(f'\n{dry_prefix}Abgeschlossen: {count_ok} gespeichert, '
                              f'{count_err} Fehler (vor KI-Abbruch).')
                        sys.exit(_EXIT_KI)

                    # Kein Rechnungsdatum → Empfangsdatum der Mail als Fallback
                    if not fields.get('InvoiceDate') and mail.received:
                        try:
                            recv_iso = str(mail.received)[:10]  # "YYYY-MM-DD"
                            y, m, d  = recv_iso.split('-')
                            fields['InvoiceDate'] = f'{d}.{m}.{y}'
                            print(f'  Warnung: Kein Rechnungsdatum gefunden — '
                                  f'Empfangsdatum der Mail wird verwendet: '
                                  f'{fields["InvoiceDate"]}', file=sys.stderr)
                        except Exception:
                            pass

                    # Keine Rechnungsnummer → aus Anhang-Dateiname extrahieren (Fallback)
                    # Erkennt lange Ziffernfolge am Dateinamen-Ende, z.B. "Invoice_10264391439.pdf"
                    if not fields.get('InvoiceNumber'):
                        m_fn = re.search(r'[_\-](\d{5,20})(?:\.[^.]+)?$', attachment.name)
                        if m_fn:
                            fields['InvoiceNumber'] = m_fn.group(1)
                            print(f'  Warnung: Keine Rechnungsnummer im PDF gefunden — '
                                  f'aus Dateiname extrahiert: {fields["InvoiceNumber"]}',
                                  file=sys.stderr)

                    # Rechnungsfilter: Pflichtfelder prüfen
                    missing_fields = [f for f in required_fields
                                      if not fields.get(f)]
                    if missing_fields:
                        print(f'  Übersprungen (keine Rechnung) — '
                              f'fehlende Felder: {", ".join(missing_fields)}')
                        continue

                    subdir   = _build_subdir(subpath_pattern, fields, fallback_dir)
                    dest_dir = os.path.join(base_dir, subdir) if subdir else base_dir
                    filename = _build_filename(filename_pattern, fields,
                                               supplier_max, invoicenr_max)
                    dest_path = _unique_path(os.path.join(dest_dir, filename + '.pdf'))

                    print(f'  {dry_prefix}-> {dest_path}')
                    print(f'    Lieferant    : {fields.get("SupplierName")  or "-"}')
                    print(f'    Rechnungsnr. : {fields.get("InvoiceNumber") or "-"}')
                    print(f'    Datum        : {fields.get("InvoiceDate")   or "-"}')
                    print(f'    Brutto       : {fields.get("GrossAmount")   or "-"}')

                    if not dry_run:
                        os.makedirs(dest_dir, exist_ok=True)
                        shutil.copy2(tmp_path, dest_path)

                    # BankingZV-Eintrag vorbereiten (auch im Dry-Run für Vorschau)
                    bzv_konto_id = ''
                    if bzv_modus and bzv_kfg.aktiv:
                        konto = _select_bzv_konto(fields, bzv_kfg)
                        if konto is None:
                            print(f'  Warnung: Kein BankingZV-Konto gefunden — '
                                  f'Eintrag übersprungen.', file=sys.stderr)
                        else:
                            bzv_konto_id = konto.konto_id
                            try:
                                bzv_e = _erstelle_bzv_eintrag(fields, konto, bzv_kfg)
                                bzv_eintraege.append(bzv_e)
                                if bzv_e['SvcLvl'] == 'SEPA':
                                    typ = 'Überweisung   '
                                elif bzv_e.get('PmtMtd') == 'DD':
                                    typ = 'Erw. Gutschrift'
                                else:
                                    typ = 'Erw. Zahlung  '
                                konto_hint = (f'  [{konto.konto_id}]'
                                              if konto.konto_id != (
                                                  bzv_kfg.standard_konto.konto_id
                                                  if bzv_kfg.standard_konto else '')
                                              else '')
                                print(f'    BankingZV  : {typ}  {bzv_e["Amt"]} EUR'
                                      f'{konto_hint}')
                            except Exception as exc:
                                print(f'  Warnung: BankingZV-Eintrag fehlgeschlagen: {exc}',
                                      file=sys.stderr)
                    elif excel_export and bzv_kfg.aktiv:
                        # Konto-ID auch ohne BZV-Export ermitteln (für Excel-Spalte N)
                        konto = _select_bzv_konto(fields, bzv_kfg)
                        if konto:
                            bzv_konto_id = konto.konto_id

                    # Excel-Export (wird auch im dry_run tatsächlich geschrieben,
                    # da -e explizit angefordert wurde; nur PDF-Save und BankingZV
                    # werden im dry_run unterdrückt)
                    if excel_export:
                        _schreibe_in_excel(fields, bzv_konto_id, excel_kfg,
                                           dry_run=False,
                                           debug=args.debug)

                    count_ok += 1

                except Exception as exc:
                    print(f'  Fehler: {exc}', file=sys.stderr)
                    count_err     += 1
                    mail_had_error = True

                finally:
                    if os.path.exists(tmp_path):
                        os.unlink(tmp_path)

            # Mail als gelesen markieren — nur wenn alle Anhänge erfolgreich
            if mark_as_read and not mail_had_error and not dry_run:
                try:
                    mail.mark_as_read()
                except Exception as exc:
                    print(f'  Warnung: Mail konnte nicht als gelesen markiert werden: {exc}',
                          file=sys.stderr)

            # Mail in Archiv-Ordner verschieben — nur wenn ohne Fehler und kein Dry-Run
            if not mail_had_error and not dry_run and mail._move_to_archive:
                try:
                    mail.move_to_archive()
                    print(f'  Mail in Archiv verschoben.')
                except Exception as exc:
                    print(f'  Warnung: Mail konnte nicht in Archiv verschoben werden: {exc}',
                          file=sys.stderr)

        # BankingZV-Export (nach Abschluss aller Mails)
        if bzv_modus and bzv_eintraege:
            _exportiere_zu_bankingzv(bzv_eintraege, bzv_kfg, bzv_modus, dry_run)

        print(f'\n{dry_prefix}Abgeschlossen: {count_ok} gespeichert, {count_err} Fehler.')

    except SystemExit:
        raise   # KI-Abbruch oder sys.exit() weitergeben (hat eigene Ausgabe)
    except Exception as exc:
        print(f'\nVerbindungsfehler: {exc}', file=sys.stderr)
        _verbindungsfehler = True
        print(f'\n{dry_prefix}Abgeschlossen: {count_ok} gespeichert, '
              f'{count_err} Fehler (Verbindungsabbruch).')
    finally:
        sys.stdout = _orig_stdout
        if _log_fh:
            _log_fh.close()

    # ── Exit-Code an aufrufende Skripte melden ────────────────────────────────
    if _verbindungsfehler:
        sys.exit(_EXIT_VERBINDUNG)
    if count_err > 0 and count_ok == 0:
        sys.exit(_EXIT_ALLE_FEHL)
    if count_err > 0:
        sys.exit(_EXIT_TEILERFOLG)
    sys.exit(_EXIT_OK)


if __name__ == '__main__':
    run()
